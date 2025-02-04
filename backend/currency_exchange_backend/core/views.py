from io import BytesIO

from rest_framework import viewsets, generics
from .models import Currency, HistoryEvent,  Shift, CustomUser
from .serializers import (
    CurrencySerializer,
    HistoryEventSerializer,
    ShiftSerializer,
    CustomUserSerializer
)
from .permissions import IsCashierOrAdmin

from rest_framework.pagination import PageNumberPagination

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.filter(is_deleted=False)
    serializer_class = CurrencySerializer
    pagination_class = StandardResultsSetPagination

    def perform_create(self, serializer):

        currency = serializer.save()

        try:
            HistoryEvent.objects.create(
                event_type='create_currency',
                user=self.request.user if self.request.user.is_authenticated else None,
                currency=currency,
                target_user=None
            )
        except Exception as e:
            print(f"Error creating history event: {e}")

    def perform_update(self, serializer):
        currency = serializer.save()
        HistoryEvent.objects.create(
            event_type='update_currency',
            user=self.request.user if self.request.user.is_authenticated else None,
            currency=currency,
            target_user=None  )

    def perform_destroy(self, instance):
        HistoryEvent.objects.create(
            event_type='delete_currency',
            user=self.request.user if self.request.user.is_authenticated else None,
            currency=instance,
            target_user=None
        )
        instance.is_deleted = True
        instance.save()
from rest_framework.exceptions import ValidationError


from rest_framework import viewsets
from rest_framework.decorators import action
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from .models import ClientOperation
from .serializers import ClientOperationSerializer

class ClientOperationViewSet(viewsets.ModelViewSet):
    queryset = ClientOperation.objects.all().order_by('-timestamp')  # Order by latest timestamp
    serializer_class = ClientOperationSerializer
    permission_classes = [IsCashierOrAdmin]

    def get_queryset(self):
        """
        Filter operations by the selected period (shift, 3 days, week).
        """
        queryset = super().get_queryset()
        period = self.request.query_params.get('period', 'shift')  # Default to 'shift'

        now = datetime.now()

        if period == 'shift':
            # Retrieve the most recent shift
            recent_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
            if recent_shift:
                queryset = queryset.filter(timestamp__gte=recent_shift.start_time)
            else:
                # Fallback to the start of the current day if no active shift is found
                start_of_day = datetime(now.year, now.month, now.day)
                queryset = queryset.filter(timestamp__gte=start_of_day)
        elif period == '3days':
            # Filter operations from the last 3 days
            three_days_ago = now - timedelta(days=3)
            queryset = queryset.filter(timestamp__gte=three_days_ago)
        elif period == 'week':
            # Filter operations from the last 7 days
            week_ago = now - timedelta(days=7)
            queryset = queryset.filter(timestamp__gte=week_ago)
        else:
            raise ValidationError("Invalid period. Use 'shift', '3days', or 'week'.")

        return queryset

    def perform_create(self, serializer):
        from rest_framework.exceptions import ValidationError
        # find active shift
        active_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
        if not active_shift or not active_shift.user:
            raise ValidationError({"detail": "No active shift or no user assigned to the active shift."})

        shift_cashier = active_shift.user

        data = serializer.validated_data
        op_type = data['operation_type']
        currency = data['currency']
        amount = data['amount']
        rate = data['exchange_rate']


        try:
            som_currency = Currency.objects.get(name="Som")
        except Currency.DoesNotExist:
            raise ValidationError({"detail": "The main currency 'Som' was not found."})

        total_som = amount * rate

        if op_type == 'buy':
            if som_currency.balance < total_som:
                raise ValidationError({"detail": "Insufficient Som balance for this purchase."})
        else:  # sell
            if currency.balance < amount:
                raise ValidationError({"detail": f"Insufficient {currency.name} balance for this sale."})

        operation = serializer.save(
            cashier_name=shift_cashier.username,  # SHIFT’s user name
            total_in_som=total_som
        )


        if op_type == 'buy':
            som_currency.balance -= total_som
            som_currency.save()
            currency.balance += amount
            currency.save()
        else:  # sell
            currency.balance -= amount
            currency.save()
            som_currency.balance += total_som
            som_currency.save()

    @action(detail=False, methods=['get'], url_path='currencies', url_name='currencies')
    def list_currencies(self, request):
        # exclude Som and also filter out deleted currencies
        currencies = Currency.objects.filter(is_deleted=False).exclude(name="Som")
        serializer = CurrencySerializer(currencies, many=True)
        return Response(serializer.data)

    @action(methods=['patch'], detail=True)
    def edit_operation(self, request, pk=None):
        """
        Edit an operation and recalculate balances.
        """
        from rest_framework.exceptions import ValidationError

        old_op = self.get_object()
        old_amount = old_op.amount
        old_total_som = old_op.total_in_som
        op_type = old_op.operation_type
        currency = old_op.currency

        try:
            som_currency = Currency.objects.get(name="Som", is_deleted=False)
        except Currency.DoesNotExist:
            raise ValidationError("The main currency 'Som' was not found or is deleted!")

        if op_type == 'buy':
            som_currency.balance += old_total_som
            currency.balance -= old_amount
        else:  # sell
            currency.balance += old_amount
            som_currency.balance -= old_total_som

        som_currency.save()
        currency.save()

        serializer = self.get_serializer(old_op, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        updated_op = serializer.save()

        new_amount = updated_op.amount
        new_rate = updated_op.exchange_rate
        new_total_som = new_amount * new_rate
        updated_op.total_in_som = new_total_som
        updated_op.edited = self.request.user.username
        updated_op.save()

        if op_type == 'buy':
            if som_currency.balance < new_total_som:
                raise ValidationError("Insufficient Som balance for this edit.")
            som_currency.balance -= new_total_som
            currency.balance += new_amount
        else:  # sell
            if currency.balance < new_amount:
                raise ValidationError(f"Insufficient {currency.name} balance for this edit.")
            currency.balance -= new_amount
            som_currency.balance += new_total_som

        som_currency.save()
        currency.save()

        return Response(self.get_serializer(updated_op).data)

    @action(detail=True, methods=['get'])
    def generate_receipt(self, request, pk=None):
        """
        Original approach: Return a downloadable PDF directly.
        Also includes 'edited_by' if operation was edited.
        """
        operation = self.get_object()

        edited_by = operation.edited if operation.edited else None

        context = {
            'operation': operation,
            'edited_by': edited_by,
        }
        html_content = render_to_string('receipt_template.html', context)

        pdf_file = BytesIO()
        pisa_status = pisa.CreatePDF(
            src=html_content,
            dest=pdf_file,
            encoding='utf-8'
        )
        if pisa_status.err:
            return Response({"error": "Error creating PDF"}, status=500)

        pdf_file.seek(0)
        filename = f"receipt_{operation.id}.pdf"
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    @action(detail=True, methods=['get'], url_path='generate_receipt_inline')
    def generate_receipt_inline(self, request, pk=None):
        """
        Return the receipt PDF as a base64-encoded string so the frontend
        can display it on-screen with 'Exit' and 'Download' buttons.
        """
        operation = self.get_object()
        edited_by = operation.edited if operation.edited else None

        context = {
            'operation': operation,
            'edited_by': edited_by,
        }

        html_content = render_to_string('receipt_template.html', context)

        pdf_file = BytesIO()
        pisa_status = pisa.CreatePDF(
            src=html_content,
            dest=pdf_file,
            encoding='utf-8'
        )
        if pisa_status.err:
            return Response({"error": "Error creating PDF"}, status=500)

        pdf_file.seek(0)
        pdf_bytes = pdf_file.read()

        import base64
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

        filename = f"receipt_{operation.id}.pdf"

        return Response({
            "pdf_base64": pdf_base64,  # The actual PDF content
            "filename": filename,  # So the frontend can name the file
        })


from django_filters.rest_framework import DjangoFilterBackend # type: ignore
...
class HistoryEventViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = HistoryEvent.objects.all().order_by('-timestamp')
    serializer_class = HistoryEventSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['event_type', 'currency__name', 'user__username']


from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
class ShiftHistoryPagination(PageNumberPagination):
    page_size = 8

class ShiftViewSet(viewsets.ModelViewSet):
    queryset = Shift.objects.all()
    serializer_class = ShiftSerializer
    @action(detail=False, methods=['get'], url_path='current_cashier')
    def current_cashier(self, request):
        """
        GET /api/shifts/current_cashier/
        Returns info about the user (cashier) of the active shift.
        """
        active_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
        if not active_shift:
            return Response({"detail": "No active shift"}, status=404)

        if not active_shift.user:
            return Response({"detail": "Active shift has no user assigned"}, status=404)

        user = active_shift.user
        return Response({
            "cashier_id": user.id,
            "cashier_username": user.username,
            "cashier_role": user.role
        })

    @action(detail=False, methods=['post'], url_path='set_cashier')
    def set_cashier(self, request):
        """
        POST /api/shifts/set_cashier/ with { "cashier_id": <some_user_id> }
        This updates the active shift's user to that user.
        """
        active_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
        if not active_shift:
            return Response({"detail": "No active shift to update"}, status=404)

        cashier_id = request.data.get("cashier_id")
        if not cashier_id:
            return Response({"detail": "cashier_id is required"}, status=400)

        try:
            new_cashier = CustomUser.objects.get(id=cashier_id, is_deleted=False)
        except CustomUser.DoesNotExist:
            return Response({"detail": "User does not exist or is deleted"}, status=404)

        old_user = active_shift.user
        active_shift.user = new_cashier
        active_shift.save()

        HistoryEvent.objects.create(
            event_type='update_user',
            user=request.user if request.user.is_authenticated else None,
            currency=None,
        )

        return Response({
            "detail": f"Active shift's user changed from {old_user} to {new_cashier.username}",
            "shift_id": active_shift.id,
            "new_cashier_id": new_cashier.id,
            "new_cashier_username": new_cashier.username,
        })

    @action(detail=False, methods=['get'], url_path='history')
    def history(self, request):
        from django.db.models import Avg, Sum
        shifts_qs = Shift.objects.all().order_by('-start_time')

        page = self.paginate_queryset(shifts_qs)
        if page is None:
            return Response([])

        results = []
        for shift in page:
            if shift.end_time:
                ops = ClientOperation.objects.filter(
                    timestamp__gte=shift.start_time,
                    timestamp__lte=shift.end_time
                )
            else:
                ops = ClientOperation.objects.filter(timestamp__gte=shift.start_time)

            ops_count = ops.count()

            buy_agg = ops.filter(operation_type='buy').aggregate(
                total_bought=Sum('amount'),
                avg_buy_rate=Avg('exchange_rate')
            )
            sell_agg = ops.filter(operation_type='sell').aggregate(
                total_sold=Sum('amount'),
                avg_sell_rate=Avg('exchange_rate')
            )

            total_bought = buy_agg['total_bought'] or 0
            avg_buy_rate = buy_agg['avg_buy_rate'] or 0
            total_sold = sell_agg['total_sold'] or 0
            avg_sell_rate = sell_agg['avg_sell_rate'] or 0

            overlap = min(total_bought, total_sold)
            profit = overlap * (avg_sell_rate - avg_buy_rate)

            raw_changes = shift.changed_balances if shift.changed_balances else []
            filtered_changes = []
            for item in raw_changes:
                old_b = item.get("old_balance")
                new_b = item.get("new_balance")
                if old_b != new_b:
                    filtered_changes.append(item)

            fmt_start = shift.start_time.strftime('%Y-%m-%d %H:%M:%S')
            if shift.end_time:
                fmt_end = shift.end_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                fmt_end = "N/A"

            results.append({
                "id": shift.id,
                "start_time": fmt_start,
                "end_time": fmt_end,
                "cashier_name": shift.user.username if shift.user else "N/A",
                "operations_count": ops_count,
                "overall_profit": round(profit, 2),
                "changed_balances": filtered_changes,
            })

        from rest_framework import serializers
        class ShiftHistoryTempSerializer(serializers.Serializer):
            id = serializers.IntegerField()
            start_time = serializers.CharField()
            end_time = serializers.CharField()
            cashier_name = serializers.CharField()
            operations_count = serializers.IntegerField()
            overall_profit = serializers.FloatField()
            changed_balances = serializers.JSONField()

        ser = ShiftHistoryTempSerializer(results, many=True)
        return self.get_paginated_response(ser.data)

    @action(methods=['post'], detail=False)  # Explicitly allow POST method
    def clear(self, request):

        active_shifts = Shift.objects.filter(end_time__isnull=True).order_by('-start_time')
        if active_shifts.exists():
            current_shift = active_shifts.first()
            current_shift.end_time = timezone.now()

            balances_data = request.data.get("balances", [])
            changes = []

            for item in balances_data:
                currency_id = item.get("currency_id")
                leftover = item.get("leftover")
                if currency_id is None or leftover is None:
                    continue

                try:
                    currency_obj = Currency.objects.get(id=currency_id)
                    if currency_obj.balance != leftover:  # Если баланс изменился
                        changes.append({
                            "currency_id": currency_id,
                            "currency_name": currency_obj.name,
                            "old_balance": float(currency_obj.balance),  # Convert to float
                            "new_balance": float(leftover),  # Convert to float
                        })
                        currency_obj.balance = Decimal(leftover)  # Use Decimal for storage
                        currency_obj.save()
                except Currency.DoesNotExist:
                    pass

            current_shift.note = f"Clear by user: {self.request.user.username}"
            current_shift.changed_balances = changes  # Сохраняем изменения
            current_shift.save()

        # Создаём новую смену
        new_shift = Shift.objects.create(user=self.request.user)
        return Response({
            "detail": "Shift cleared and new shift opened",
            "new_shift_id": new_shift.id,
            "force_logout": True
        })


from rest_framework.permissions import IsAuthenticated
from .serializers import HistoryEventSerializer

class InternalHistoryPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class InternalHistoryAPIView(generics.ListAPIView):
    queryset = HistoryEvent.objects.all().order_by('-timestamp')  # Adjust ordering as needed
    serializer_class = HistoryEventSerializer
    pagination_class = InternalHistoryPagination
    permission_classes = [IsAuthenticated]


from .permissions import IsAdminOrReadOnly
class CustomUserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.filter(is_deleted=False)
    serializer_class = CustomUserSerializer
    permission_classes = [IsAdminOrReadOnly]
    pagination_class = StandardResultsSetPagination

    def perform_create(self, serializer):
        # Save the new user instance
        user = serializer.save()
        # Create a HistoryEvent for user creation
        HistoryEvent.objects.create(
            event_type='create_user',
            user=self.request.user if self.request.user.is_authenticated else None,
            currency=None,  # No currency associated with user creation
            target_user=user  # The user being created
        )

    def perform_destroy(self, instance):
        # Create a HistoryEvent for user deletion
        HistoryEvent.objects.create(
            event_type='delete_user',
            user=self.request.user if self.request.user.is_authenticated else None,
            currency=None,  # No currency associated with user deletion
            target_user=instance  # The user being deleted
        )
        # Soft-delete the user
        instance.is_deleted = True
        instance.save()



from rest_framework.views import APIView
from .permissions import IsCashierOrAdmin

class AnalyticsView(APIView):
    permission_classes = [IsCashierOrAdmin]

    def get(self, request):
        from datetime import timedelta
        period = request.GET.get('period', 'today')

        now = timezone.now()

        if period == 'week':
            start = now - timedelta(days=7)
        elif period == 'month':
            start = now - timedelta(days=30)
        elif period == '3days':
            start = now - timedelta(days=3)
        elif period == 'shift':
            # последняя смена
            last_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
            if last_shift:
                start = last_shift.start_time
            else:
                start = now - timedelta(days=3)
        else:
            # 'today' — полночь текущего дня
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)

        # Фильтруем операции за период
        ops = ClientOperation.objects.filter(timestamp__gte=start, timestamp__lte=now)

        # Считаем аналитику
        # например, как было
        som_balance = 0
        try:
            som_balance = Currency.objects.get(name="Som").balance
        except Currency.DoesNotExist:
            pass

        results = []
        total_profit = 0
        all_currencies = Currency.objects.filter(is_deleted=False)



        for cur in all_currencies:
            if cur.name == "Som":
                continue
            buy_ops = ops.filter(operation_type='buy', currency=cur)
            sell_ops = ops.filter(operation_type='sell', currency=cur)
            cur_balance = cur.balance

            buy_count = buy_ops.aggregate(total_buy_amount=Sum('amount'))['total_buy_amount'] or 0
            sell_count = sell_ops.aggregate(total_sell_amount=Sum('amount'))['total_sell_amount'] or 0
            avg_buy_rate = buy_ops.aggregate(Avg('exchange_rate'))['exchange_rate__avg'] or 0
            avg_sell_rate = sell_ops.aggregate(Avg('exchange_rate'))['exchange_rate__avg'] or 0
            usable_count = min(buy_count, sell_count)  # берем минимум
            profit = usable_count * (avg_sell_rate - avg_buy_rate)

            total_profit += profit

            results.append({
                "currency": cur.name,
                "balance": round(cur_balance, 2),
                "buy_count": buy_count,
                "avg_buy_rate": round(avg_buy_rate, 4),
                "sell_count": sell_count,
                "avg_sell_rate": round(avg_sell_rate, 4),
                "profit": round(profit, 2),
            })

        data = {
            "period": period,
            "start_time": start,
            "end_time": now,
            "som_balance": som_balance,
            "total_profit": round(total_profit, 2),
            "details": results
        }
        return Response(data)
from .models import HistoryEvent, Shift
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from rest_framework.views import APIView
from datetime import  datetime
from django.utils.timezone import now as timezone_now
from decimal import Decimal
from django.db.models import Q


class ExportAnalyticsExcel(APIView):
    """
    Export analytics data to an Excel file with no authentication required.
    """
    authentication_classes = []
    permission_classes = []

    def get(self, request, *args, **kwargs):
        period = request.GET.get('period', 'today')
        now = timezone_now()

        if period == 'week':
            start = now - timedelta(days=7)
        elif period == 'month':
            start = now - timedelta(days=30)
        elif period == '3days':
            start = now - timedelta(days=3)
        elif period == 'shift':
            last_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
            start = last_shift.start_time if last_shift else now - timedelta(days=3)
        else:
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)

        ops = ClientOperation.objects.filter(timestamp__gte=start, timestamp__lte=now)

        results = []
        total_profit = Decimal('0.00')
        all_currencies = Currency.objects.exclude(name="Som").filter(is_deleted=False)

        for cur in all_currencies:
            buy_ops = ops.filter(operation_type='buy', currency=cur)
            sell_ops = ops.filter(operation_type='sell', currency=cur)

            buy_count = buy_ops.aggregate(total_buy_amount=Sum('amount'))['total_buy_amount'] or Decimal('0.00')
            sell_count = sell_ops.aggregate(total_sell_amount=Sum('amount'))['total_sell_amount'] or Decimal('0.00')
            avg_buy_rate = buy_ops.aggregate(Avg('exchange_rate'))['exchange_rate__avg'] or Decimal('0.00')
            avg_sell_rate = sell_ops.aggregate(Avg('exchange_rate'))['exchange_rate__avg'] or Decimal('0.00')

            profit = min(buy_count, sell_count) * (avg_sell_rate - avg_buy_rate)
            total_profit += profit

            results.append({
                "currency": cur.name,
                "buy_count": float(buy_count),
                "avg_buy_rate": round(float(avg_buy_rate), 4),
                "sell_count": float(sell_count),
                "avg_sell_rate": round(float(avg_sell_rate), 4),
                "profit": round(float(profit), 2),
            })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics"

        headers = ["Currency", "Buy Count", "Avg Buy Rate", "Sell Count", "Avg Sell Rate", "Profit"]
        header_font = Font(bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_num, row in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=row['currency'])
            ws.cell(row=row_num, column=2, value=row['buy_count'])
            ws.cell(row=row_num, column=3, value=row['avg_buy_rate'])
            ws.cell(row=row_num, column=4, value=row['sell_count'])
            ws.cell(row=row_num, column=5, value=row['avg_sell_rate'])
            ws.cell(row=row_num, column=6, value=row['profit'])

        total_row = len(results) + 2
        ws.cell(row=total_row, column=5, value="Total Profit:")
        ws.cell(row=total_row, column=6, value=round(float(total_profit), 2))
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=6).font = Font(bold=True)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"analytics_{period}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportEventExcel(APIView):

    authentication_classes = []
    permission_classes = []

    def get(self, request, *args, **kwargs):
        period = request.GET.get('period', '3days')
        now = timezone_now()

        if period == 'week':
            start = now - timedelta(days=7)
        elif period == 'month':
            start = now - timedelta(days=30)
        elif period == 'shift':
            last_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
            start = last_shift.start_time if last_shift else now - timedelta(days=3)
        elif period == '3days':
            start = now - timedelta(days=3)
        else:
            start = now - timedelta(days=3)

        events = HistoryEvent.objects.filter(
            timestamp__gte=start, timestamp__lte=now
        ).select_related('user', 'currency', 'target_user').exclude(
            Q(user__is_deleted=True) | Q(currency__is_deleted=True)
        ).order_by('-timestamp')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"

        headers = ["ID", "Event Type", "User", "Currency", "Target User", "Timestamp"]
        header_font = Font(bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_num, ev in enumerate(events, 2):
            ws.cell(row=row_num, column=1, value=ev.id)
            ws.cell(row=row_num, column=2, value=ev.get_event_type_display())
            ws.cell(row=row_num, column=3, value=ev.user.username if ev.user else "")
            ws.cell(row=row_num, column=4, value=ev.currency.name if ev.currency else "")
            ws.cell(row=row_num, column=5, value=ev.target_user.username if ev.target_user else "")
            ws.cell(row=row_num, column=6, value=ev.timestamp.strftime('%Y-%m-%d %H:%M:%S'))

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"events_{period}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportOperationExcel(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, *args, **kwargs):
        period = request.GET.get('period', '3days')
        ops = self.filter_by_period(period)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Operations"

        headers = ["ID", "Type", "Currency", "Cashier", "Amount", "Exchange", "Total Som", "Timestamp"]
        header_font = Font(bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_num, op in enumerate(ops, 2):
            ws.cell(row=row_num, column=1, value=op.id)
            ws.cell(row=row_num, column=2, value=op.operation_type)
            ws.cell(row=row_num, column=3, value=op.currency.name)
            ws.cell(row=row_num, column=4, value=op.cashier_name)
            ws.cell(row=row_num, column=5, value=float(op.amount))
            ws.cell(row=row_num, column=6, value=float(op.exchange_rate))
            ws.cell(row=row_num, column=7, value=float(op.total_in_som))
            ws.cell(row=row_num, column=8, value=op.timestamp.strftime('%Y-%m-%d %H:%M:%S'))

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"operations_{period}_{timezone_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def filter_by_period(self, period):
        now = timezone_now()
        if period == 'week':
            start = now - timedelta(days=7)
        elif period == 'month':
            start = now - timedelta(days=30)
        elif period == 'shift':
            last_shift = Shift.objects.filter(end_time__isnull=True).order_by('-start_time').first()
            start = last_shift.start_time if last_shift else now - timedelta(days=3)
        elif period == '3days':
            start = now - timedelta(days=3)
        else:
            start = now - timedelta(days=3)

        return ClientOperation.objects.filter(timestamp__gte=start).order_by('-timestamp')


from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum, Avg, Count
from django.db.models.functions import TruncHour
from django.utils import timezone
from datetime import timedelta

from .models import ClientOperation, Currency
from .permissions import IsCashierOrAdmin


class AdvancedAnalyticsView(APIView):
    permission_classes = [IsCashierOrAdmin]

    def get(self, request):
        period = request.GET.get('period', 'week')
        now = timezone.now()

        if period == 'month':
            start = now - timedelta(days=30)
        elif period == '3days':
            start = now - timedelta(days=3)
        else:  # default 'week'
            start = now - timedelta(days=7)

        ops = ClientOperation.objects.filter(timestamp__gte=start, timestamp__lte=now)


        results = []
        total_profit = 0

        currencies = Currency.objects.exclude(name="Som")

        for cur in currencies:
            buy_ops = ops.filter(operation_type='buy', currency=cur)
            sell_ops = ops.filter(operation_type='sell', currency=cur)

            buy_count = buy_ops.aggregate(total_buy_amount=Sum('amount'))['total_buy_amount'] or 0
            sell_count = sell_ops.aggregate(total_sell_amount=Sum('amount'))['total_sell_amount'] or 0

            avg_buy = buy_ops.aggregate(Avg('exchange_rate'))['exchange_rate__avg'] or 0
            avg_sell = sell_ops.aggregate(Avg('exchange_rate'))['exchange_rate__avg'] or 0

            usable_count = min(buy_count, sell_count)
            profit = usable_count * (avg_sell - avg_buy)
            total_profit += profit

            results.append({
                "currency": cur.name,
                "buy_count": buy_count,
                "sell_count": sell_count,
                "avg_buy_rate": round(avg_buy, 4),
                "avg_sell_rate": round(avg_sell, 4),
                "profit": round(profit, 2),
            })

        peak_hours_qs = (
            ops.annotate(hour=TruncHour('timestamp'))
            .values('hour')
            .annotate(operation_count=Count('id'))
            .order_by('-operation_count')[:2]
        )
        peak_hours = [
            {
                "hour": entry['hour'].strftime('%H:%M'),
                "operation_count": entry['operation_count']
            }
            for entry in peak_hours_qs
        ]

        total_transactions = ops.count()

        total_buys = ops.filter(operation_type='buy').count()

        total_sells = ops.filter(operation_type='sell').count()

        average_profit_per_transaction = round(total_profit / total_transactions, 2) if total_transactions else 0

        overall_operations = total_buys + total_sells

        data = {
            "period": period,
            "start_time": start,
            "end_time": now,
            "total_profit": round(total_profit, 2),
            "total_transactions": total_transactions,
            "total_buys": total_buys,
            "total_sells": total_sells,
            "average_profit_per_transaction": average_profit_per_transaction,
            "overall_operations": overall_operations,
            "peak_hours": peak_hours,
            "details": results
        }

        return Response(data)
